"""Export to XLSX and CSV."""
import io
import csv
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_xlsx(area: str, overall: dict, by_type: list, listings: list, insights: list, furnishing: list) -> bytes:
    wb = Workbook()

    # Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = f"Temu Properti — Market Analysis · {area}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A4"] = "Metric"
    ws1["B4"] = "Value"
    for cell in (ws1["A4"], ws1["B4"]):
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="002FA7")
        cell.font = Font(bold=True, color="FFFFFF")
    rows = [
        ("Listing Count", overall.get("listing_count", 0)),
        ("Average Price (RM)", overall.get("average_price", 0)),
        ("Median Price (RM)", overall.get("median_price", 0)),
        ("Mode Price (RM)", overall.get("mode_price", 0)),
        ("Fair Price (RM)", overall.get("fair_price", 0)),
        ("Average Size (sqft)", overall.get("average_size", 0)),
        ("Rent per sqft (RM)", overall.get("rent_per_sqft", 0)),
        ("Min Price (RM)", overall.get("min_price", 0)),
        ("Max Price (RM)", overall.get("max_price", 0)),
    ]
    for i, (k, v) in enumerate(rows, start=5):
        ws1[f"A{i}"] = k
        ws1[f"B{i}"] = v
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # By unit type
    row_start = len(rows) + 7
    ws1[f"A{row_start}"] = "By Unit Type"
    ws1[f"A{row_start}"].font = Font(bold=True, size=12)
    headers = ["Unit Type", "Count", "Avg Price", "Median", "Mode", "Fair Price", "Avg Size (sqft)"]
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=row_start + 1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    for i, t in enumerate(by_type, start=row_start + 2):
        ws1.cell(row=i, column=1, value=t["unit_type"])
        ws1.cell(row=i, column=2, value=t["listing_count"])
        ws1.cell(row=i, column=3, value=t["average_price"])
        ws1.cell(row=i, column=4, value=t["median_price"])
        ws1.cell(row=i, column=5, value=t["mode_price"])
        ws1.cell(row=i, column=6, value=t["fair_price"])
        ws1.cell(row=i, column=7, value=t["average_size"])

    # Insights sheet
    ws2 = wb.create_sheet("Insights")
    ws2["A1"] = "Automated Market Insights"
    ws2["A1"].font = Font(bold=True, size=14)
    for i, insight in enumerate(insights, start=3):
        ws2[f"A{i}"] = f"• {insight}"
        ws2[f"A{i}"].alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 100

    # Furnishing sheet
    ws3 = wb.create_sheet("Furnishing")
    ws3["A1"] = "Furnishing"
    ws3["B1"] = "Count"
    ws3["C1"] = "Percentage"
    for c in (ws3["A1"], ws3["B1"], ws3["C1"]):
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    for i, f in enumerate(furnishing, start=2):
        ws3.cell(row=i, column=1, value=f["name"])
        ws3.cell(row=i, column=2, value=f["value"])
        ws3.cell(row=i, column=3, value=f"{f['percentage']}%")

    # Listings sheet
    ws4 = wb.create_sheet("Listings")
    cols = ["Title", "Property", "Area", "Bedroom Type", "Monthly Rent (RM)", "Annual Rent (RM)", "Size (sqft)", "Furnishing", "URL"]
    for col, h in enumerate(cols, start=1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    for i, li in enumerate(listings, start=2):
        ws4.cell(row=i, column=1, value=li.get("title"))
        ws4.cell(row=i, column=2, value=li.get("property_name"))
        ws4.cell(row=i, column=3, value=li.get("area_name"))
        ws4.cell(row=i, column=4, value=li.get("bedroom_type"))
        ws4.cell(row=i, column=5, value=li.get("monthly_rent"))
        ws4.cell(row=i, column=6, value=li.get("annual_rent"))
        ws4.cell(row=i, column=7, value=li.get("size_sqft"))
        ws4.cell(row=i, column=8, value=li.get("furnishing_status"))
        ws4.cell(row=i, column=9, value=li.get("listing_url"))
    for col in range(1, 10):
        ws4.column_dimensions[chr(64 + col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(listings: list) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Title", "Property", "Area", "Bedroom Type", "Monthly Rent (RM)", "Annual Rent (RM)", "Size (sqft)", "Furnishing", "URL"])
    for li in listings:
        w.writerow([
            li.get("title", ""),
            li.get("property_name", ""),
            li.get("area_name", ""),
            li.get("bedroom_type", ""),
            li.get("monthly_rent", ""),
            li.get("annual_rent", ""),
            li.get("size_sqft", ""),
            li.get("furnishing_status", ""),
            li.get("listing_url", ""),
        ])
    return buf.getvalue().encode("utf-8")
